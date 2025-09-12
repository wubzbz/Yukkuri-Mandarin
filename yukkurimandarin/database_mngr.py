# 数据库管理操作。

from pathlib import Path
from typing import List, Tuple, Optional
import csv
from datetime import datetime

from yukkurimandarin.database import Database

# 可选组件
try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


class DatabaseManager:
    """拼音数据库管理类"""

    # 默认文件IO路径
    DEFAULT_FILE_DIR = Path.cwd()
    DEFAULT_XLSX_PATH = DEFAULT_FILE_DIR / "yinjie_table.xlsx"
    DEFAULT_CSV_PATH = DEFAULT_FILE_DIR / "yinjie_table.csv"

    def __init__(self, db_path: Optional[str] = None):
        if db_path is None:
            self.db_path = Database.DEFAULT_DB_PATH
        else:
            self.db_path = Path(db_path) # lazy initialization


    def _get_width(self, s: str) -> int:
        """计算字符串的显示宽度，全宽字符计为2，半宽字符计为1"""
        width = 0
        for char in s:
            # 全宽字符范围：
            # - CJK统一汉字：\u4e00-\u9fff
            # - 全角标点：\u3000-\u303f
            # - 全角英数符号：\uff01-\uff5e
            # - 全角平假名：\u3040-\u309f
            # - 全角片假名：\u30a0-\u30ff（不包含半角片假名\uff61-\uff9f）
            code = ord(char)
            if 0x4e00 <= code <= 0x9fff or \
            0x3000 <= code <= 0x303f or \
            0x3040 <= code <= 0x309f or \
            0x30a0 <= code <= 0x30ff or \
            0xff01 <= code <= 0xff5e:
                width += 2
            else:
                width += 1
        return width


    def _report_result(self, result: dict) -> None:
        """输出操作结果"""
        # 提取结果数据
        operation = result.get("操作", "未知操作")
        success = result.get("结果", False)
        #rows_affected = result.get("影响行数", 0)
        data = result.get("数据", [])
        info = result.get("信息", [])
        #time_taken = result.get("用时", 0)
        # 打印时间
        current_time = datetime.now()
        print(f"@ {current_time.strftime('%Y-%m-%d %H:%M:%S')} -- 操作:{operation} {'成功' if success else '失败'}")
        if data and isinstance(data[0], tuple):
            # 表头
            headers = ["音节", "声调", "平假名"]
            # 按列计算最大宽度
            column_widths = []
            header_widths = []
            data_widths = []
            for h in headers:
                hw = self._get_width(h)
                column_widths.append(hw)
                header_widths.append(hw)
            for row in data:
                row_widths = []
                for i, item in enumerate(row):
                    w = self._get_width(item)
                    row_widths.append(w)
                    if w > column_widths[i]:
                        column_widths[i] = w
                data_widths.append(row_widths)
                    
            # 打印分隔线
            separator = "|-" + "-|-".join("-" * w for w in column_widths) + "-|"
            print(separator)
            # 打印表头
            header_row = "| " + " | ".join(f"{header:^{cw - (hw - len(header))}}" for header, hw, cw in zip(headers, header_widths, column_widths)) + " |"
            print(header_row)
            print(separator)
            # 打印数据行
            for row, rw in zip(data, data_widths):
                row_str = "| " + " | ".join(f"{cell:^{cw - (dw - len(cell))}}" for cell, dw, cw in zip(row, rw, column_widths)) + " |"
                print(row_str)
            print(separator)
        if info and isinstance(info[0], str):
            for i in info:
                print(i) 
        #print(f"影响{rows_affected}行. 用时{time_taken}秒. ")


    def add_pinyin(self, yinjie: str, tone: str, hiragana: str, report: bool = True) -> bool:
        """
        向拼音数据库中增加拼音
        
        Args:
            yinjie: 音节（不含声调的拼音）
            tone: 前一字、本字、后一字的声调
            hiragana: 对应的平假名拟音（含音声记号）
            report: 是否打印操作结果

        Returns:
            操作是否成功
        """
        if len(tone) != 3 or any(t not in "012345" for t in tone) or tone[1] == "0":
            if report:
                result = {"操作": "增加拼音数据",
                        "结果": False,
                        "信息": [f"发生错误：声调 {tone} 格式错误！"]}
                self._report_result(result)
            return False
        try:
            self.db = Database(self.db_path)
            self.db.insert_entry(yinjie, tone, hiragana)
            self.db.close()
            if report:
                result = {"操作": "增加拼音数据",
                        "结果": True,
                        "数据": [(yinjie, tone, hiragana)]}
                self._report_result(result)
            return True
        except Exception as e:
            if report:
                result = {"操作": "增加拼音数据",
                        "结果": False,
                        "信息": [f"发生错误：{e}"]}
                self._report_result(result)
            return False


    def search_by_pinyin(self, yinjie: str, tone: str = "***", report: bool = True) -> List[Tuple[str, str, str]]:
        """在拼音数据库中查找拼音对应的平假名拟音

        Args:
            yinjie: 音节（不含声调的拼音）
            tone: 前一字、本字、后一字的声调（通配符*）
            report: 是否打印操作结果
        
        Returns:
            对应的平假名拟音（含音声记号）
        """
        if any(t not in "012345*" for t in tone) or tone[1] == "0":
            if report:
                result = {"操作": "查询拼音数据",
                        "结果": False,
                        "信息": [f"发生错误：声调 {tone} 格式错误！"]}
                self._report_result(result)
            return []
        try:
            self.db = Database(self.db_path)
            record = []
            if tone == "***":
                record += self.db.query_by_yinjie(yinjie)
            else:
                first_digits = "012345" if tone[0] == "*" else [tone[0]]
                second_digits = "12345" if tone[1] == "*" else [tone[1]]
                third_digits = "012345" if tone[2] == "*" else [tone[2]]

                for d1 in first_digits:
                    for d2 in second_digits:
                        for d3 in third_digits:
                            record += self.db.query_by_pinyin(yinjie, f"{d1}{d2}{d3}")
            self.db.close()
            if report:
                result = {"操作": "查询拼音数据",
                        "结果": True,
                        "数据": record}
                self._report_result(result)
            return record
        except Exception as e:
            if report:
                result = {"操作": "查询拼音数据",
                        "结果": False,
                        "信息": [f"发生错误：{e}"]}
                self._report_result(result)
            return []
    

    def serial_search(self, serial: List[Tuple[str, str]], default: str = "") -> List[str]:
        """拼音序列搜索（长连接）
        
        Args:
            serial: 拼音序列，由(yinjie, tone)组成的列表
            default: 若搜索无结果的返回值

            * `keep`: 保留音节
            * 其他字符：替换为该字符
        
        Returns:
            转换结果
        """
        if not serial:
            return []
        # 查询序列
        self.db = Database(self.db_path)
        hiragana_list = self.db.query_batch(serial, default)
        self.db.close()
        # 检查长度
        if len(serial) != len(hiragana_list):
            raise ValueError("序列查询与结果不匹配。")
        return hiragana_list


    def delete_pinyin(self, yinjie: str, tone: str, report: bool = True) -> bool:
        """删除拼音数据
        
        Args:
            yinjie: 音节（不含声调的拼音）
            tone: 前一字、本字、后一字的声调（通配符*）
            report: 是否打印操作结果

        Returns:
            操作是否成功
            """
        if len(tone) != 3 or any(t not in "012345*" for t in tone) or tone[1] == "0":
            if report:
                result = {"操作": "删除拼音数据",
                        "结果": False,
                        "信息": [f"发生错误：声调 {tone} 格式错误！"]}
                self._report_result(result)
            return False
        try:
            self.db = Database(self.db_path)
            record = []
            if tone == "***":
                record += self.db.query_by_yinjie(yinjie)
                self.db.delete_by_yinjie(yinjie)
            else:
                first_digits = "012345" if tone[0] == "*" else [tone[0]]
                second_digits = "12345" if tone[1] == "*" else [tone[1]]
                third_digits = "012345" if tone[2] == "*" else [tone[2]]

                for d1 in first_digits:
                    for d2 in second_digits:
                        for d3 in third_digits:
                            record += self.db.query_by_pinyin(yinjie, f"{d1}{d2}{d3}")
                            self.db.delete_by_pinyin(yinjie, f"{d1}{d2}{d3}")
            self.db.close()
            if report:
                result = {"操作": "删除拼音数据",
                        "结果": True,
                        "数据": record}
                self._report_result(result)
            return True
        except Exception as e:
            if report:
                result = {"操作": "查询拼音数据",
                        "结果": False,
                        "信息": [f"发生错误：{e}"]}
                self._report_result(result)
            return False


    def export_to_csv(self, file_path: Optional[str] = None, report: bool = True) -> bool:
        """导出数据到csv文件
            
        Args:
            file_path: 导出的文件路径
            report: 是否打印操作结果
        
        Returns:
            操作是否成功
        """
        if file_path is None:
            filepath = self.DEFAULT_CSV_PATH
        else:
            filepath = Path(file_path)
        try:
            # 检查路径
            if filepath.suffix.lower() != ".csv":
                filepath.with_suffix(".csv")
            # 添加表头
            rows = [("拼音", "声调", "平假名")]
            # 获取所有数据
            self.db = Database(self.db_path)
            data = self.db.query_all()
            self.db.close()
            # 填充数据
            rows += data
            # 保存文件
            with open(filepath, mode="w", newline="", encoding="utf-8-sig") as file:
                writer = csv.writer(file)
                writer.writerows(rows)
            if report:
                result = {"操作": "导出拼音数据库到csv",
                        "结果": True,
                        "信息": [f"成功导出到{filepath}"]}
                self._report_result(result)
            return True
        except Exception as e:
            if report:
                result = {"操作": "导出拼音数据库到csv",
                        "结果": False,
                        "信息": [f"发生错误：{e}"]}
                self._report_result(result)
            return False


    def import_from_csv(self, file_path: Optional[str] = None, report: bool = True) -> bool:
        """从csv文件导入数据
            
        Args:
            file_path: 导入的文件路径
            report: 是否打印操作结果
        
        Returns:
            操作是否成功
        """
        if file_path is None:
            filepath = self.DEFAULT_CSV_PATH
        else:
            filepath = Path(file_path)
        try:
            # 检查路径
            if filepath.suffix.lower() != ".csv":
                if report:
                    result = {"操作": "从csv导入拼音数据库",
                            "结果": False,
                            "信息": ["发生错误：文件格式不正确。需要.csv文件。"]}
                    self._report_result(result)
                return False
            # 读取数据
            rows = []
            errors = []
            with open(filepath, "r", newline="", encoding="utf-8-sig") as csvfile:
                reader = csv.reader(csvfile)
                # 验证每行是否包含3列
                for row_num, row in enumerate(reader, start=1):
                    if len(row) != 3:
                        errors.append(f"第{row_num}行包含{len(row)}列数据，期望为3列")
                        continue
                    rows.append(tuple(row))
            # 验证表头
            expected_headers = ("拼音", "声调", "平假名")
            if not rows or rows[0] != expected_headers:
                if report:
                    result = {"操作": "从csv导入拼音数据库",
                            "结果": False,
                            "信息": [f"发生错误：表头不匹配{expected_headers}。"]}
                    self._report_result(result)
                return False
                
            # 验证数据格式并收集有效数据
            valid_entries = []
            errors = []
                
            for row_idx in range(1, len(rows)):
                pinyin = rows[row_idx][0]
                tone = rows[row_idx][1]
                hiragana = rows[row_idx][2]
                    
                # 检查必填字段
                if not pinyin or tone is None or not hiragana:
                    errors.append(f"第 {row_idx} 行: 缺少必填字段")
                    continue
                # 检查声调是否为3位数
                if len(tone) != 3:
                    errors.append(f"第 {row_idx} 行: 声调{tone}必须是3位数字")
                    continue
                if any(t not in "012345" for t in tone) or tone[1] == "0":
                    errors.append(f"第 {row_idx} 行: 声调{tone}格式错误")
                    continue
                
                valid_entries.append((pinyin, tone, hiragana))
            
            # 插入有效数据
            if valid_entries:
                self.db = Database(self.db_path)
                self.db.insert_batch(valid_entries)
                self.db.close()

            errors.append(f"导入完成。成功: {len(valid_entries)}, 失败: {len(errors)}")
            if report:
                result =  {"操作": "从csv导入拼音数据库",
                           "结果": len(errors) == 1,
                           "信息": errors}
                self._report_result(result)
            return len(errors) == 1
        
        except Exception as e:
            if report:
                result = {"操作": "从csv导入拼音数据库",
                        "结果": False,
                        "信息": [f"发生错误：{e}"]}
                self._report_result(result)
            return False


    def export_to_excel(self, file_path: Optional[str] = None, report: bool = True) -> bool:
        """导出数据到Excel文件
            
        Args:
            file_path: 导出的文件路径
            report: 是否打印操作结果
        
        Returns:
            操作是否成功
        """
        if not _HAS_OPENPYXL:
            if report:
                result = {"操作": "导出拼音数据库到excel表格",
                        "结果": False,
                        "信息": ["未安装openpyxl，该功能不可用。\n请尝试通过 pip install openpyxl 进行安装。 或使用fill_csv方法代替。"]}
                self._report_result(result)
            return False
        if file_path is None:
            filepath = self.DEFAULT_XLSX_PATH
        else:
            filepath = Path(file_path)
        try:
            # 检查路径
            if filepath.suffix.lower() != ".xlsx":
                filepath.with_suffix(".xlsx")
            # 创建工作簿和工作表
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            if not worksheet:
                worksheet = workbook.create_sheet(title="拼音数据")
                workbook.active = worksheet
            worksheet.title = "拼音数据"
            # 添加表头
            worksheet["A1"] = "拼音"
            worksheet["B1"] = "声调"
            worksheet["C1"] = "平假名"
            # 获取所有数据
            self.db = Database(self.db_path)
            data = self.db.query_all()
            self.db.close()
            # 填充数据
            for row_idx, row_data in enumerate(data, 2):
                pinyin, tone, hiragana = row_data
                worksheet[f"A{row_idx}"] = pinyin
                worksheet[f"B{row_idx}"] = tone
                worksheet[f"C{row_idx}"] = hiragana
            # 保存文件
            workbook.save(filepath)
            if report:
                result = {"操作": "导出拼音数据库到excel表格",
                        "结果": True,
                        "信息": [f"成功导出到{filepath}"]}
                self._report_result(result)
            return True
        except Exception as e:
            if report:
                result = {"操作": "导出拼音数据库到excel表格",
                        "结果": False,
                        "信息": [f"发生错误：{e}"]}
                self._report_result(result)
            return False
        

    def import_from_excel(self, file_path: Optional[str] = None, report: bool = True) -> bool:
        """从Excel文件导入数据
            
        Args:
            file_path: 导入的文件路径
            report: 是否打印操作结果
        
        Returns:
            操作是否成功
        """
        if not _HAS_OPENPYXL:
            if report:
                result = {"操作": "从excel表格导入拼音数据库",
                        "结果": False,
                        "信息": ["未安装openpyxl，该功能不可用。\n请尝试通过 pip install openpyxl 进行安装。 或使用fill_csv方法代替。"]}
                self._report_result(result)
            return False
        if file_path is None:
            filepath = self.DEFAULT_XLSX_PATH
        else:
            filepath = Path(file_path)
        try:
            # 检查路径
            if filepath.suffix.lower() != ".xlsx":
                if report:
                    result = {"操作": "从excel表格导入拼音数据库",
                            "结果": False,
                            "信息": ["发生错误：文件格式不正确。需要.xlsx文件。"]}
                    self._report_result(result)
                return False
            # 打开工作簿
            workbook = openpyxl.load_workbook(filepath)
            worksheet = workbook.active
            if not worksheet:
                if report:
                    result = {"操作": "从excel表格导入拼音数据库",
                            "结果": False,
                            "信息": [f"发生错误：缺少工作表。"]}
                    self._report_result(result)
                return False
            # 验证表头
            expected_headers = ["拼音", "声调", "平假名"]
            actual_headers = [worksheet.cell(row=1, column=i).value for i in range(1, 4)]
                
            if actual_headers != expected_headers:
                if report:
                    result = {"操作": "从excel表格导入拼音数据库",
                            "结果": False,
                            "信息": [f"发生错误：表头不匹配{expected_headers}。"]}
                    self._report_result(result)
                return False
                
            # 验证数据格式并收集有效数据
            valid_entries = []
            errors = []
                
            for row_idx in range(2, worksheet.max_row + 1):
                pinyin = worksheet.cell(row=row_idx, column=1).value
                tone = worksheet.cell(row=row_idx, column=2).value
                if not isinstance(tone, str):
                    tone = str(tone)
                hiragana = worksheet.cell(row=row_idx, column=3).value
                    
                # 检查必填字段
                if not pinyin or tone is None or not hiragana:
                    errors.append(f"第 {row_idx} 行: 缺少必填字段")
                    continue
                # 检查声调是否为3位数
                if len(tone) != 3:
                    errors.append(f"第 {row_idx} 行: 声调必须是3位数字")
                    continue
                if any(t not in "012345" for t in tone) or tone[1] == "0":
                    errors.append(f"第 {row_idx} 行: 声调格式错误")
                    continue
                
                valid_entries.append((pinyin, tone, hiragana))
            
            # 插入有效数据
            if valid_entries:
                self.db = Database(self.db_path)
                self.db.insert_batch(valid_entries)
                self.db.close()

            errors.append(f"导入完成。成功: {len(valid_entries)}, 失败: {len(errors)}")
            if report:
                result =  {"操作": "从excel表格导入拼音数据库",
                           "结果": len(errors) == 1,
                           "信息": errors}
                self._report_result(result)
            return len(errors) == 1
        
        except Exception as e:
            if report:
                result = {"操作": "从excel表格导入拼音数据库",
                        "结果": False,
                        "信息": [f"发生错误：{e}"]}
                self._report_result(result)
            return False